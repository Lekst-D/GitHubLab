from PyQt5 import uic 
from PyQt5.QtWidgets import QMainWindow,QTableWidgetItem,QFileDialog 
from ui.CellsMushroom import Ui_CellsMushroom, Ui_CellsMushroomEdit
from ui.KeeperMushroom import Ui_KeeperMushroom, Ui_KeeperMushroomEdit
from ui.Mushroom import Ui_Mushroom, Ui_MushroomEdit
import openpyxl 

class Ui_Main(QMainWindow):

    def __init__(self,connection):
        self.count = 1000
        self.listPrice = list()
        self.listValue = list()
        super().__init__()
        uic.loadUi('ui\Main.ui', self)  

        self.con = connection
        self.cursor = self.con.cursor()

        #События с продажей
        self.pushButton.clicked.connect(self.openNewCells)
        self.pushButton_2.clicked.connect(self.openEditCells)
        self.pushButton_3.clicked.connect(self.deleteCells)

        #События с хранилищем
        self.pushButton_9.clicked.connect(self.openNewKeeper)
        self.pushButton_11.clicked.connect(self.openEditKeeper)
        self.pushButton_10.clicked.connect(self.deleteKeeper)

        #События с грибами
        self.pushButton_5.clicked.connect(self.openNewMushroom)
        self.pushButton_6.clicked.connect(self.openEditMushroom)
        self.pushButton_7.clicked.connect(self.DeleteMushroom)

        # Заполнение таблиц на форме
        self.setRowsMushrooms()
        self.setRowsKeeperMushrooms()
        self.setRowsCellsMushrooms()

        #Создание отчетов
        self.pushButton_4.clicked.connect(self.createOtchetCells)
        self.pushButton_12.clicked.connect(self.createOtchetKeeper)
        self.pushButton_8.clicked.connect(self.createOtchetMushrooms)

    def saveFile(self):
        filename, ok = QFileDialog.getSaveFileName(self,
                             "Сохранить файл",
                             ".",
                             "*.xlsx")
        resultList = [filename, ok]
        return resultList

#Создание отчетов
    def createOtchetCells(self):   
        wb = openpyxl.Workbook() 
        sheet = wb.active 

        results = self.cursor.execute("SELECT cells_mushroom.Id, mushroom.Name,CountMushroom,DateRealize,MoneyAbandon FROM cells_mushroom join  mushroom on mushroom.Id  = cells_mushroom.IdMushroom ").fetchall()
        i=1

        c1 = sheet.cell(row = i, column = 1) 
        c1.value = "Ид"
        c1 = sheet.cell(row = i, column = 2) 
        c1.value = "Имя гриба"
        c1 = sheet.cell(row = i, column = 3) 
        c1.value = "Количество"
        c1 = sheet.cell(row = i, column = 4) 
        c1.value = "Дата продажи"
        c1 = sheet.cell(row = i, column = 5) 
        c1.value = "Выручка"

        i=2

        for result in results:

            c1 = sheet.cell(row = i, column = 1) 
            c1.value = str(result[0])
            c1 = sheet.cell(row = i, column = 2) 
            c1.value = str(result[1])
            c1 = sheet.cell(row = i, column = 3) 
            c1.value = str(result[2])
            c1 = sheet.cell(row = i, column = 4) 
            c1.value = str(result[3])
            c1 = sheet.cell(row = i, column = 5) 
            c1.value = str(result[4])

            i+=1
        
        trip = (self.saveFile()[0]).replace('/', '\\\\') 

        wb.save(trip)
    
    def createOtchetKeeper(self): 
        wb = openpyxl.Workbook() 
        sheet = wb.active 

        results = self.cursor.execute("SELECT store_mushroom.Id,mushroom.Name,Count,DateRemain,TimeLife FROM store_mushroom join  mushroom on mushroom.Id  = store_mushroom.IdMushroom ").fetchall()
        i=1

        c1 = sheet.cell(row = i, column = 1) 
        c1.value = "Ид"
        c1 = sheet.cell(row = i, column = 2) 
        c1.value = "Имя гриба"
        c1 = sheet.cell(row = i, column = 3) 
        c1.value = "Количество"
        c1 = sheet.cell(row = i, column = 4) 
        c1.value = "Начало хранения"
        c1 = sheet.cell(row = i, column = 5) 
        c1.value = "Время хранения(дней)"

        i=2

        for result in results:

            c1 = sheet.cell(row = i, column = 1) 
            c1.value = str(result[0])
            c1 = sheet.cell(row = i, column = 2) 
            c1.value = str(result[1])
            c1 = sheet.cell(row = i, column = 3) 
            c1.value = str(result[2])
            c1 = sheet.cell(row = i, column = 4) 
            c1.value = str(result[3])
            c1 = sheet.cell(row = i, column = 5) 
            c1.value = str(result[4])

            i+=1
        
        trip = (self.saveFile()[0]).replace('/', '\\\\') 

        wb.save(trip)

    def createOtchetMushrooms(self): 
        wb = openpyxl.Workbook() 
        sheet = wb.active 
        
        results = self.cursor.execute("SELECT mushroom.Id, mushroom.Name,Description,TimeLife,CountMushroom,DateCost FROM mushroom join  price_mushroom on mushroom.Id  = price_mushroom.IdMushroom  ").fetchall()
        
        i=1

        c1 = sheet.cell(row = i, column = 1) 
        c1.value = "Ид"
        c1 = sheet.cell(row = i, column = 2) 
        c1.value = "Имя гриба"
        c1 = sheet.cell(row = i, column = 3) 
        c1.value = "Описание"
        c1 = sheet.cell(row = i, column = 4) 
        c1.value = "Время хранения"
        c1 = sheet.cell(row = i, column = 5) 
        c1.value = "Стоимость"
        c1 = sheet.cell(row = i, column = 6) 
        c1.value = "День установки цены"

        i=2

        for result in results:

            c1 = sheet.cell(row = i, column = 1) 
            c1.value = str(result[0])
            c1 = sheet.cell(row = i, column = 2) 
            c1.value = str(result[1])
            c1 = sheet.cell(row = i, column = 3) 
            c1.value = str(result[2])
            c1 = sheet.cell(row = i, column = 4) 
            c1.value = str(result[3])
            c1 = sheet.cell(row = i, column = 5) 
            c1.value = str(result[4])
            c1 = sheet.cell(row = i, column = 6) 
            c1.value = str(result[5])

            i+=1
        
        trip = (self.saveFile()[0]).replace('/', '\\\\') 

        wb.save(trip) 

#Заполнение таблиц на форме
    def setRowsMushrooms(self):
        self.con.commit()

        #Таблица грибов
        results = self.cursor.execute("SELECT mushroom.Id, mushroom.Name,Description,TimeLife,CountMushroom,DateCost FROM mushroom join  price_mushroom on mushroom.Id  = price_mushroom.IdMushroom  ").fetchall()
        
        self.tableWidget.removeRow(0)
        self.tableWidget.setColumnCount(6)
        self.tableWidget.setHorizontalHeaderLabels(['ИД','Имя гриба','Описание','Время жизни','Стоимость','День установки цены']) 

        i = 0

        print(results)

        for result in results:
            self.tableWidget.setRowCount(i+1)

            #qdate = QtCore.QDate.fromString(result[5], 'dd/MM/yyyy')

            self.tableWidget.setItem(i,0,QTableWidgetItem(str(result[0])))
            self.tableWidget.setItem(i,1,QTableWidgetItem(str(result[1])))
            self.tableWidget.setItem(i,2,QTableWidgetItem(str(result[2])))
            self.tableWidget.setItem(i,3,QTableWidgetItem(str(result[3])))
            self.tableWidget.setItem(i,4,QTableWidgetItem(str(result[4])))
            self.tableWidget.setItem(i,5,QTableWidgetItem(str(result[5])))

            i+=1

    def setRowsKeeperMushrooms(self):
        self.con.commit()

        #Таблица грибов хранилище
        results = self.cursor.execute("SELECT store_mushroom.Id,mushroom.Name,Count,DateRemain,TimeLife FROM store_mushroom join  mushroom on mushroom.Id  = store_mushroom.IdMushroom ").fetchall()
        
        self.tableWidget_2.removeRow(0)
        self.tableWidget_2.setColumnCount(5)
        self.tableWidget_2.setHorizontalHeaderLabels(['ИД','Имя гриба','Количество','Время начало хранения','Время жизни']) 

        i = 0

        print(results)

        for result in results:
            self.tableWidget_2.setRowCount(i+1)

            #qdate = QtCore.QDate.fromString(result[5], 'dd/MM/yyyy')

            self.tableWidget_2.setItem(i,0,QTableWidgetItem(str(result[0])))
            self.tableWidget_2.setItem(i,1,QTableWidgetItem(str(result[1])))
            self.tableWidget_2.setItem(i,2,QTableWidgetItem(str(result[2])))
            self.tableWidget_2.setItem(i,3,QTableWidgetItem(str(result[3])))
            self.tableWidget_2.setItem(i,4,QTableWidgetItem(str(result[4])))

            i+=1

    def setRowsCellsMushrooms(self):
        self.con.commit()

        #Таблица грибов продажа
        results = self.cursor.execute("SELECT cells_mushroom.Id, mushroom.Name,CountMushroom,DateRealize,MoneyAbandon FROM cells_mushroom join  mushroom on mushroom.Id  = cells_mushroom.IdMushroom ").fetchall()
        
        self.tableWidget_3.removeRow(0)
        self.tableWidget_3.setColumnCount(5)
        self.tableWidget_3.setHorizontalHeaderLabels(['ИД','Имя гриба','Количество','Дата продажи','Получено']) 

        i = 0

        print(results)

        for result in results:
            self.tableWidget_3.setRowCount(i+1)

            #qdate = QtCore.QDate.fromString(result[5], 'dd/MM/yyyy')

            self.tableWidget_3.setItem(i,0,QTableWidgetItem(str(result[0])))
            self.tableWidget_3.setItem(i,1,QTableWidgetItem(str(result[1])))
            self.tableWidget_3.setItem(i,2,QTableWidgetItem(str(result[2])))
            self.tableWidget_3.setItem(i,3,QTableWidgetItem(str(result[3])))
            self.tableWidget_3.setItem(i,4,QTableWidgetItem(str(result[4])))

            i+=1

#Работа с продажей грибов
    def openNewCells(self):
        self.another_form = Ui_CellsMushroom(self,self.con)
        self.another_form.show()

    def deleteCells(self):
        row = (self.tableWidget_3.currentRow())
        if(row!=-1):
            id = self.tableWidget_3.item(row, 0)
            id = id.text()
            self.cursor.execute(f"DELETE from cells_mushroom where Id = {id};")
            self.setRowsCellsMushrooms()

    def openEditCells(self):
        row = (self.tableWidget_3.currentRow())
        if(row!=-1):
            id = self.tableWidget_3.item(row, 0)
            id = id.text()
            self.another_form = Ui_CellsMushroomEdit(self,self.con,id)
            self.another_form.show()

#Работа с хранилищем грибов
    def openNewKeeper(self):
        self.another_form = Ui_KeeperMushroom(self,self.con)
        self.another_form.show()

    def deleteKeeper(self):
        row = (self.tableWidget_2.currentRow())
        if(row!=-1):
            id = self.tableWidget_2.item(row, 0)
            id = id.text()
            self.cursor.execute(f"DELETE from store_mushroom where Id = {id};")
            self.setRowsKeeperMushrooms()
            
    def openEditKeeper(self):
        row = (self.tableWidget_2.currentRow())
        if(row!=-1):
            id = self.tableWidget_2.item(row, 0)
            id = id.text()
            self.another_form = Ui_KeeperMushroomEdit(self,self.con,id)
            self.another_form.show()

#Работа с грибами
    def openNewMushroom(self):
        self.another_form = Ui_Mushroom(self,self.con)
        self.another_form.show()
    
    def openEditMushroom(self):
        row = (self.tableWidget.currentRow())
        if(row!=-1):
            id = self.tableWidget.item(row, 0)
            id = id.text()
            self.another_form = Ui_MushroomEdit(self,self.con,id)
            self.another_form.show()

    def DeleteMushroom(self):
        row = (self.tableWidget.currentRow())
        if(row!=-1):
            id = self.tableWidget.item(row, 0)
            id = id.text()
            self.cursor.execute(f"DELETE from price_mushroom where IdMushroom = {id};")
            self.cursor.execute(f"DELETE from mushroom where Id = {id};")
            self.setRowsMushrooms()
    
